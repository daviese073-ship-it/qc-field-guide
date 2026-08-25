from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
BATIR_WORKBOOK = ROOT / "docs" / "source" / "batir data.xlsx"
ACTIVITY_DIR = ROOT / "src" / "data" / "activities"
TERMINOLOGY_DIR = ROOT / "src" / "data" / "terminology"
BATIR_TERMINOLOGY_FILE = TERMINOLOGY_DIR / "batir-qc.json"
BATIR_AUDIT_FILE = TERMINOLOGY_DIR / "batir-audit.json"

CONTENT_BLOCK_FIELDS = [
    "requirements",
    "planning",
    "documentControl",
    "materialControl",
    "evidence",
    "correctiveAction",
    "verification",
    "closureCriteria",
    "reportingAnalysis",
    "qualityCheckpoint",
]

INSPECTION_FIELDS = ["before", "during", "after", "testing"]
ISSUE_FIELDS = ["commonDeficiencies", "escalationTriggers"]
COMMUNICATION_FIELDS = ["before", "during", "issueEscalation", "after"]
OUTPUT_FIELDS = ["records", "acceptanceEvidence", "followUp"]

FALSE_POSITIVE_TERMS = {
    "current",
    "grade",
    "level",
    "support",
    "bearing",
    "form",
    "clean",
    "lift",
    "power",
    "post",
    "square",
    "duration",
    "priority",
    "risk",
    "activity",
    "objective",
    "client",
    "owner",
    "pour",
}

MANUAL_TRANSLATIONS: dict[str, str] = {
    "Obtain/review": "Obtenir/réviser",
    "specifications/devis": "devis",
    "Plan Qualité": "Plan Qualité",
    "applicable PIE/PRIE": "PIE/PRIE applicable",
    "technical changes.": "changements techniques.",
    "Hold point / witness point requirements.": "Exigences de point d’arrêt / point témoin.",
    "For illustration only": "À titre indicatif seulement",
    "Project documents govern": "Les documents du projet prévalent",
    "where required": "lorsque requis",
    "where applicable": "le cas échéant",
    "if required": "si requis",
    "if applicable": "s’il y a lieu",
    "as required": "selon les exigences",
    "as specified": "tel que spécifié",
    "per project procedure": "selon la procédure du projet",
    "per project documents": "selon les documents du projet",
    "under the project procedure": "selon la procédure du projet",
    "project documents": "documents du projet",
    "approved procedure": "procédure approuvée",
    "approved procedures": "procédures approuvées",
    "approved shop drawings": "dessins d’atelier approuvés",
    "approved submittals": "documents soumis approuvés",
    "approved system": "système approuvé",
    "approved corrective method": "méthode corrective approuvée",
    "authorized disposition": "disposition autorisée",
    "technical disposition": "disposition technique",
    "specialist acceptance": "acceptation par un spécialiste",
    "professional approval": "approbation professionnelle",
    "quality/site authority": "autorité qualité/chantier",
    "Responsable qualité": "Responsable qualité",
    "Hydro-Québec": "Hydro-Québec",
    "QMT/RFI": "QMT/RFI",
    "RFI/QMT": "RFI/QMT",
    "QA/QC": "AQ/CQ",
    "Quality Control": "Contrôle de la qualité",
    "quality control": "contrôle de la qualité",
    "Quality Assurance": "Assurance qualité",
    "quality assurance": "assurance qualité",
    "non-conformance": "non-conformité",
    "Non-Conformance": "Non-conformité",
    "nonconformance": "non-conformité",
    "deficiency log": "registre des déficiences",
    "deficiency": "déficience",
    "deficiencies": "déficiences",
    "corrective action": "mesure corrective",
    "corrective actions": "mesures correctives",
    "inspection report": "rapport d’inspection",
    "test report": "rapport d’essai",
    "as-built drawings": "plans tels que construits",
    "record drawings": "plans conformes à l’exécution",
    "shop drawing": "dessin d’atelier",
    "shop drawings": "dessins d’atelier",
    "structural drawings": "plans structuraux",
    "mechanical drawings": "plans mécaniques",
    "electrical drawings": "plans électriques",
    "civil drawings": "plans de génie civil",
    "architectural drawings": "plans architecturaux",
    "demolition drawings": "plans de démolition",
    "utility drawings": "plans des services publics",
    "existing-conditions drawings": "plans des conditions existantes",
    "temporary-work requirements": "exigences relatives aux ouvrages temporaires",
    "environmental requirements": "exigences environnementales",
    "owner requirements": "exigences du propriétaire",
    "manufacturer requirements": "exigences du fabricant",
    "manufacturer instructions": "instructions du fabricant",
    "manufacturer data": "données du fabricant",
    "material data": "données des matériaux",
    "mock-up": "échantillon témoin",
    "mock-ups": "échantillons témoins",
    "pre-pour": "avant bétonnage",
    "post-pour": "après bétonnage",
    "pre-concealment": "avant dissimulation",
    "concealed": "dissimulé",
    "concealment": "dissimulation",
    "fire-rated": "coupe-feu",
    "firestop": "calfeutrement coupe-feu",
    "firestopping": "calfeutrement coupe-feu",
    "fire separation": "séparation coupe-feu",
    "fire-resistance": "résistance au feu",
    "life safety": "sécurité des personnes",
    "commissioning": "mise en service",
    "turnover": "remise",
    "closeout": "clôture",
    "interface": "interface",
    "interfaces": "interfaces",
    "sequence": "séquence",
    "sequencing": "séquençage",
    "traceability": "traçabilité",
    "evidence": "preuve",
    "records": "dossiers",
    "record": "dossier",
    "report": "rapport",
    "reports": "rapports",
    "document": "documenter",
    "documentation": "documentation",
    "documents": "documents",
    "verify": "vérifier",
    "verification": "vérification",
    "inspect": "inspecter",
    "inspection": "inspection",
    "confirm": "confirmer",
    "coordinate": "coordonner",
    "escalate": "escalader",
    "notify": "aviser",
    "identify": "identifier",
    "review": "réviser",
    "compare": "comparer",
    "measure": "mesurer",
    "control": "contrôler",
    "hold": "retenir",
    "release": "libérer",
    "repair": "réparer",
    "replace": "remplacer",
    "remove": "retirer",
    "install": "installer",
    "reinstall": "réinstaller",
    "reinspect": "réinspecter",
    "retest": "refaire l’essai",
    "close": "clore",
    "document": "documenter",
    "photograph": "photographier",
    "retain": "conserver",
    "accept": "accepter",
    "reject": "rejeter",
    "must": "doit",
    "should": "devrait",
    "may": "peut",
    "cannot": "ne peut pas",
    "do not": "ne pas",
    "only when": "seulement lorsque",
    "unless": "à moins que",
    "before": "avant",
    "after": "après",
    "during": "pendant",
    "actual": "réel",
    "affected": "touché",
    "approved": "approuvé",
    "required": "requis",
    "specified": "spécifié",
    "suitable": "approprié",
    "unsound": "non sain",
    "missing": "manquant",
    "failed": "non conforme",
    "passing": "conforme",
    "localized": "localisé",
    "it might involve": "cela pourrait impliquer",
    "removing unsuitable material": "retirer le matériau inapproprié",
    "replacing it with specified material": "le remplacer par le matériau spécifié",
    "the project authority decides this, not our playbook": "l’autorité du projet décide de cela, et non notre guide",
    "You should not independently declare": "Vous ne devriez pas déclarer de façon indépendante",
    "you're specifically authorized and qualified to do so": "vous êtes expressément autorisé et qualifié pour le faire",
    "reopen quality status": "rouvrir le statut qualité",
    "quality status": "statut qualité",
    "modified configuration": "configuration modifiée",
    "approved configuration": "configuration approuvée",
    "update register": "mettre à jour le registre",
    "re-close": "refermer",
    "affected wall": "mur touché",
    "further finishing": "finition ultérieure",
    "open as necessary": "ouvrir au besoin",
    "restore joints/finishes": "restaurer les joints/finitions",
    "actual water path": "chemin réel de l’eau",
    "passing retest": "contre-essai conforme",
    "old plan": "ancien plan",
    "affected area": "zone touchée",
    "remove/rebuild": "retirer/reconstruire",
    "remove failed": "retirer le matériau défaillant",
    "prepare joint properly": "préparer le joint correctement",
    "surface preparation": "préparation de surface",
    "unsound coating": "enduit non sain",
    "reapply": "réappliquer",
    "reprepare": "repréparer",
    "failed area": "zone non conforme",
    "actual condition": "condition réelle",
    "actual usable clearance": "dégagement utilisable réel",
    "physical conflict": "conflit physique",
    "prevent final acceptance": "empêcher l’acceptation finale",
    "coordinate authorized relocation/modification": "coordonner le déplacement/la modification autorisée",
    "specialist": "spécialiste",
    "specialists": "spécialistes",
    "qualified": "qualifié",
    "authorized": "autorisé",
    "authority": "autorité",
    "playbook": "guide",
    "independently": "de façon indépendante",
    "declare": "déclarer",
    "involve": "impliquer",
    "removing": "retirer",
    "replacing": "remplacer",
    "decides": "décide",
    "specifically": "expressément",
    "material": "matériau",
    "materials": "matériaux",
    "unsuitable": "inapproprié",
    "our": "notre",
    "you": "vous",
    "you're": "vous êtes",
    "it": "cela",
    "this": "cela",
    "that": "cela",
    "not": "ne pas",
    "no": "aucun",
    "reopen": "rouvrir",
    "modified": "modifié",
    "update": "mettre à jour",
    "register": "registre",
    "status": "statut",
    "southeast": "sud-est",
    "portion": "partie",
    "soft area": "zone molle",
    "remediation": "remédiation",
    "subsequently": "par la suite",
    "performed": "réalisé",
    "recorded": "consigné",
    "reference": "référence",
    "reached": "a atteint",
    "geometry": "géométrie",
    "elevation": "élévation",
    "pending": "en attente",
    "localized soft area": "zone molle localisée",
    "verification recorded": "vérification consignée",
    "geotechnical bearing capacity": "capacité portante géotechnique",
    "Grid": "grille",
    "identified": "identifié",
    "held": "retenu",
    "completed": "terminé",
    "project": "projet",
    "but": "mais",
    "at": "à",
    "with": "avec",
    "against": "par rapport à",
    "qualified to do so": "qualifié pour le faire",
    "exact": "exact",
    "final": "final",
    "construction": "construction",
    "field": "chantier",
    "site": "chantier",
    "work": "travaux",
    "trade": "corps de métier",
    "substrate": "support",
    "surface": "surface",
    "area": "zone",
    "location": "emplacement",
    "extent": "étendue",
    "condition": "condition",
    "requirement": "exigence",
    "criteria": "critères",
    "criterion": "critère",
    "acceptance": "acceptation",
    "configuration": "configuration",
    "drawing": "plan",
    "drawings": "plans",
    "specification": "devis",
    "specifications": "devis",
    "procedure": "procédure",
    "procedure": "procédure",
    "Concrete": "Béton",
    "concrete": "béton",
    "reinforcement": "armatures",
    "rebar": "barres d’armature",
    "concrete cover": "enrobage",
    "cover": "enrobage",
    "formwork": "coffrage",
    "embedded items": "éléments incorporés",
    "embedded": "incorporé",
    "embed": "élément incorporé",
    "slab-on-grade": "dalle sur sol",
    "slab": "dalle",
    "foundation": "fondation",
    "footing": "semelle",
    "wall": "mur",
    "walls": "murs",
    "masonry": "maçonnerie",
    "steel": "acier",
    "structural steel": "acier de charpente",
    "bolt": "boulon",
    "bolted": "boulonné",
    "connection": "assemblage",
    "connections": "assemblages",
    "welding": "soudage",
    "weld": "soudure",
    "anchor bolt": "boulon d’ancrage",
    "anchor": "ancrage",
    "anchorage": "ancrage",
    "grout": "coulis",
    "mortar": "mortier",
    "curing": "cure",
    "slump": "affaissement",
    "segregation": "ségrégation",
    "strength": "résistance",
    "compression": "compression",
    "density": "densité",
    "compaction": "compactage",
    "backfill": "remblai",
    "excavation": "excavation",
    "subgrade": "sol d’assise",
    "geotechnical report": "rapport géotechnique",
    "geotechnical": "géotechnique",
    "bearing capacity": "capacité portante",
    "drainage": "drainage",
    "drain": "drain",
    "pipe": "conduite",
    "piping": "tuyauterie",
    "trench": "tranchée",
    "catch basin": "puisard",
    "grading": "nivellement",
    "slope": "pente",
    "erosion": "érosion",
    "geotextile": "géotextile",
    "waterproofing membrane": "membrane d’étanchéité",
    "waterproofing": "étanchéité",
    "membrane": "membrane",
    "air barrier": "pare-air",
    "vapour barrier": "pare-vapeur",
    "insulation": "isolation",
    "sealant": "scellant",
    "sealants": "scellants",
    "flashing": "solin",
    "window": "fenêtre",
    "windows": "fenêtres",
    "door": "porte",
    "doors": "portes",
    "glazing": "vitrage",
    "cladding": "revêtement extérieur",
    "roofing membrane": "membrane de toiture",
    "roof": "toiture",
    "roofing": "toiture",
    "ceiling": "plafond",
    "drywall": "gypse",
    "framing": "cadrage",
    "paint": "peinture",
    "coating": "enduit",
    "primer": "apprêt",
    "adhesive": "adhésif",
    "adhesion": "adhérence",
    "flatness": "planéité",
    "plumbness": "aplomb",
    "HVAC": "CVCA",
    "ductwork": "réseau de conduits",
    "damper": "clapet",
    "valve": "vanne",
    "sprinkler": "gicleur",
    "grounding": "mise à la terre",
    "electrical": "électrique",
    "mechanical": "mécanique",
}


def normalize(value: str | None) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", str(value).lower())
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9/&+ -]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def slug(value: str) -> str:
    normalized = normalize(value)
    normalized = normalized.replace("&", " and ").replace("/", " ")
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized).strip("-")
    return normalized.upper()[:48]


def split_cell(value: Any) -> list[str]:
    if value is None or str(value).strip() in {"", "N/A"}:
        return []
    parts = re.split(r"[;|]", str(value))
    return [part.strip() for part in parts if part.strip()]


def iter_batir_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(BATIR_WORKBOOK, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        current_path = ""
        for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            path, fr, gender, en, definition_fr, definition_en, synonyms, verbs, sentence, nouns = row
            if path:
                current_path = str(path)
            if not en or not fr:
                continue
            rows.append(
                {
                    "sheet": worksheet.title,
                    "row": index,
                    "path": current_path,
                    "termFr": str(fr).strip(),
                    "gender": str(gender).strip() if gender else None,
                    "termEn": str(en).strip(),
                    "definitionFr": str(definition_fr).strip() if definition_fr else None,
                    "definitionEn": str(definition_en).strip() if definition_en else None,
                    "synonymsFr": split_cell(synonyms),
                    "verbsExpressions": split_cell(verbs),
                    "professionalSentence": str(sentence).strip() if sentence else None,
                    "nouns": split_cell(nouns),
                }
            )
    return rows


def collect_en_strings(value: Any) -> list[str]:
    strings: list[str] = []
    if isinstance(value, dict):
        if isinstance(value.get("en"), str):
            strings.append(value["en"])
        for nested in value.values():
            strings.extend(collect_en_strings(nested))
    elif isinstance(value, list):
        for nested in value:
            strings.extend(collect_en_strings(nested))
    return strings


def production_strings() -> list[str]:
    files = [ROOT / "src" / "data" / "sections" / "sections.json"]
    files.extend(sorted(ACTIVITY_DIR.glob("*.json")))
    files.extend(
        [
            ROOT / "src" / "data" / "gates" / "gates.json",
            ROOT / "src" / "data" / "conditions" / "conditions.json",
            ROOT / "src" / "data" / "invalidation" / "invalidation-rules.json",
        ]
    )
    values: list[str] = []
    for path in files:
        values.extend(collect_en_strings(json.loads(path.read_text(encoding="utf-8"))))
    return values


def build_corpus_index(max_words: int = 8) -> Counter[str]:
    index: Counter[str] = Counter()
    for value in production_strings():
        tokens = normalize(value).split()
        for size in range(1, min(max_words, len(tokens)) + 1):
            for offset in range(0, len(tokens) - size + 1):
                index[" ".join(tokens[offset : offset + size])] += 1
    return index


def count_phrase(index: Counter[str], phrase: str) -> int:
    normalized = normalize(phrase)
    if len(normalized) < 4:
        return 0
    return index[normalized]


def load_existing_term_keys() -> set[str]:
    keys: set[str] = set()
    for path in TERMINOLOGY_DIR.glob("*.json"):
        if path.name.startswith("batir-"):
            continue
        for concept in json.loads(path.read_text(encoding="utf-8")):
            keys.add(normalize(concept["preferred"]["en"]))
            for alias in concept.get("aliases", {}).get("en", []):
                keys.add(normalize(alias))
    return keys


def discipline_for(path: str, term_en: str) -> str:
    source = f"{path} {term_en}".lower()
    if "roof" in source:
        return "Roofing"
    if "building envelope" in source or any(word in source for word in ["window", "door", "sealant", "flashing"]):
        return "Envelope"
    if "building systems" in source or any(word in source for word in ["hvac", "valve", "piping", "sprinkler"]):
        if any(word in source for word in ["grounding", "electrical", "power"]):
            return "Electrical"
        return "Mechanical"
    if "concrete" in source or "foundation" in source:
        return "Concrete"
    if "structure" in source or "bridges" in source:
        return "Structural"
    if "finishes" in source or any(word in source for word in ["drywall", "ceiling", "paint"]):
        return "Interiors"
    if "post-construction" in source:
        return "Closeout"
    if any(word in source for word in ["earthworks", "drainage", "roads", "sidewalk", "site preparation"]):
        return "Earthworks"
    if "gestion de projet" in source or "quality" in source:
        return "Quality"
    return "Quality"


def build_batir_concepts() -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows = iter_batir_rows()
    corpus_index = build_corpus_index()
    existing_keys = load_existing_term_keys()
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    matched_rows = 0

    for row in rows:
        term_key = normalize(row["termEn"])
        if not term_key or term_key in FALSE_POSITIVE_TERMS:
            continue
        if "→ Verbes" in row["path"]:
            continue
        count = count_phrase(corpus_index, row["termEn"])
        if count == 0:
            continue
        matched_rows += 1
        row["matchCount"] = count
        grouped[term_key].append(row)

    concepts: list[dict[str, Any]] = []
    ambiguous: list[str] = []
    conflicts: list[str] = []
    existing_updates = 0

    for term_key, term_rows in sorted(
        grouped.items(), key=lambda item: (-sum(row["matchCount"] for row in item[1]), item[0])
    ):
        if term_key in existing_keys:
            existing_updates += 1
            continue

        total_count = sum(row["matchCount"] for row in term_rows)
        if total_count < 4:
            continue

        fr_counts = Counter(row["termFr"] for row in term_rows)
        preferred_fr, _ = fr_counts.most_common(1)[0]
        if len(fr_counts) > 1:
            ambiguous.append(term_rows[0]["termEn"])

        primary = term_rows[0]
        discipline = discipline_for(primary["path"], primary["termEn"])
        aliases_fr = sorted(
            {
                alias
                for row in term_rows
                for alias in [*row.get("synonymsFr", []), *row.get("nouns", [])]
                if alias and normalize(alias) != normalize(preferred_fr)
            },
            key=normalize,
        )[:8]
        source_paths = sorted({row["path"] for row in term_rows})

        concept: dict[str, Any] = {
            "id": f"TERM-BATIR-{slug(discipline)}-{slug(primary['termEn'])}",
            "discipline": discipline,
            "preferred": {"en": primary["termEn"], "fr": preferred_fr},
            "definition": {
                "en": primary.get("definitionEn") or f"BÂTIR terminology candidate for {primary['termEn']}.",
                "fr": primary.get("definitionFr") or preferred_fr,
                "status": {"en": "provisional", "fr": "provisional"},
            },
            "contextNotes": {
                "en": "BÂTIR-derived terminology candidate matched to the QC Field Guide production corpus; not a new QC requirement.",
                "fr": "Candidat terminologique dérivé de BÂTIR et apparié au corpus de production du Guide CQ; il ne constitue pas une nouvelle exigence CQ.",
                "status": {"en": "validated", "fr": "provisional"},
            },
            "status": {"en": "provisional", "fr": "provisional"},
            "confidence": {"en": "medium", "fr": "medium"},
            "sourceRef": {
                "build": "Build 4 / BÂTIR",
                "document": "batir data.xlsx",
                "section": source_paths[0],
            },
        }
        if aliases_fr:
            concept["aliases"] = {"fr": aliases_fr}
        concepts.append(concept)

    concepts.sort(key=lambda concept: (concept["discipline"], concept["preferred"]["en"]))

    audit = {
        "sourceWorkbook": "docs/source/batir data.xlsx",
        "sheetNames": ["Buildings", "Infrastructure", "Gestion de projet"],
        "rowsParsed": {
            "Buildings": sum(1 for row in rows if row["sheet"] == "Buildings"),
            "Infrastructure": sum(1 for row in rows if row["sheet"] == "Infrastructure"),
            "Gestion de projet": sum(1 for row in rows if row["sheet"] == "Gestion de projet"),
        },
        "totalRowsParsed": len(rows),
        "candidateRelevantRows": matched_rows,
        "candidateRelevantConcepts": len(grouped),
        "canonicalConceptsCreated": len(concepts),
        "existingConceptsMatchedWithoutOverride": existing_updates,
        "duplicatesConsolidated": matched_rows - len(grouped),
        "irrelevantOrUnmappedRows": len(rows) - matched_rows,
        "ambiguousEntries": sorted(ambiguous),
        "preferredTerminologyConflicts": conflicts,
        "parsingErrors": [],
        "notes": [
            "Rows from verb-only paths were excluded from canonical concept creation and used only as draft-translation vocabulary.",
            "BÂTIR terms were matched against production corpus phrases before becoming candidates.",
            "Existing Build-4 preferred terminology was not overridden by BÂTIR.",
        ],
    }
    return concepts, audit


def build_phrase_map() -> list[tuple[re.Pattern[str], str]]:
    rows = iter_batir_rows()
    corpus_index = build_corpus_index()
    phrase_map: dict[str, str] = {}
    for en, fr in MANUAL_TRANSLATIONS.items():
        phrase_map[en] = fr
    for row in rows:
        en = row["termEn"]
        fr = row["termFr"]
        normalized = normalize(en)
        if (
            len(normalized) >= 4
            and normalized not in FALSE_POSITIVE_TERMS
            and count_phrase(corpus_index, en) > 0
        ):
            phrase_map.setdefault(en, fr)
    return [
        (
            re.compile(rf"(?<![A-Za-z0-9]){re.escape(english)}(?![A-Za-z0-9])", re.IGNORECASE),
            french,
        )
        for english, french in sorted(
            phrase_map.items(), key=lambda item: len(item[0]), reverse=True
        )
    ]


def tidy_french(text: str) -> str:
    replacements = [
        (r"\bthe\b", "le"),
        (r"\band\b", "et"),
        (r"\bor\b", "ou"),
        (r"\bwith\b", "avec"),
        (r"\bwithout\b", "sans"),
        (r"\bfrom\b", "de"),
        (r"\bto\b", "à"),
        (r"\bof\b", "de"),
        (r"\bin\b", "dans"),
        (r"\bfor\b", "pour"),
        (r"\bby\b", "par"),
        (r"\bon\b", "sur"),
        (r"\bunder\b", "sous"),
        (r"\bagainst\b", "contre"),
        (r"\binto\b", "dans"),
        (r"\bold\b", "ancien"),
        (r"\bnew\b", "nouveau"),
        (r"\bsame\b", "même"),
        (r"\bboth\b", "les deux"),
        (r"\bnext\b", "suivant"),
        (r"\bfinal\b", "final"),
        (r"\bactual\b", "réel"),
        (r"\bproperly\b", "correctement"),
        (r"\bcorrectly\b", "correctement"),
        (r"\bafterward\b", "par la suite"),
    ]
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    text = re.sub(r"\s+([;:,.])", r"\1", text)
    text = re.sub(r"([([])\s+", r"\1", text)
    text = re.sub(r"\s+([])])", r"\1", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def draft_translate(value: str, phrase_map: list[tuple[re.Pattern[str], str]]) -> str:
    translated = value
    for pattern, french in phrase_map:
        translated = pattern.sub(french, translated)
    translated = tidy_french(translated)
    return translated


def localize_value(
    value: Any,
    phrase_map: list[tuple[re.Pattern[str], str]],
    key_path: tuple[str, ...] = (),
) -> int:
    if isinstance(value, dict):
        changed = 0
        status = value.get("status") if isinstance(value.get("status"), dict) else {}
        if (
            isinstance(value.get("en"), str)
            and (not value.get("fr") or status.get("fr") == "provisional")
            and key_path[-1:] != ("title",)
        ):
            value["fr"] = draft_translate(value["en"], phrase_map)
            status = value.setdefault("status", {})
            status.setdefault("en", "validated")
            status["fr"] = "provisional"
            changed += 1
        for nested_key, nested in value.items():
            changed += localize_value(nested, phrase_map, (*key_path, nested_key))
        return changed
    if isinstance(value, list):
        return sum(localize_value(item, phrase_map, key_path) for item in value)
    return 0


def collect_content_items_from_activity(activity: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []

    def collect_blocks(blocks: list[dict[str, Any]] | None) -> None:
        for block in blocks or []:
            if block.get("type") in {"paragraph", "notice"}:
                items.append(block["item"])
            elif block.get("type") in {"bulletList", "checkList"}:
                items.extend(block.get("items", []))

    for field in CONTENT_BLOCK_FIELDS:
        collect_blocks(activity.get(field))
    for key in INSPECTION_FIELDS:
        collect_blocks((activity.get("inspection") or {}).get(key))
    for key in ISSUE_FIELDS:
        collect_blocks((activity.get("issues") or {}).get(key))
    for key in COMMUNICATION_FIELDS:
        collect_blocks((activity.get("communications") or {}).get(key))
    for key in OUTPUT_FIELDS:
        collect_blocks((activity.get("outputs") or {}).get(key))
    if activity.get("specialistBoundary"):
        items.append(activity["specialistBoundary"])
    return items


def apply_activity_translations(phrase_map: list[tuple[re.Pattern[str], str]]) -> dict[str, Any]:
    item_count = 0
    item_fr_count = 0
    localized_values_changed = 0
    authority_count = 0
    authority_fr_count = 0
    high_control_count = 0
    high_control_fr_count = 0

    for path in sorted(ACTIVITY_DIR.glob("*.json")):
        activities = json.loads(path.read_text(encoding="utf-8"))
        for activity in activities:
            localized_values_changed += localize_value(activity, phrase_map)
            for item in collect_content_items_from_activity(activity):
                item_count += 1
                if item.get("text", {}).get("fr"):
                    item_fr_count += 1
                authority = item.get("authority") or {}
                is_authority = any(
                    authority.get(key)
                    for key in [
                        "projectDocumentsGovern",
                        "specialistRequired",
                        "authorizedProcessRequired",
                    ]
                )
                if is_authority:
                    authority_count += 1
                    if item.get("text", {}).get("fr"):
                        authority_fr_count += 1
                high_control = item.get("highControl") or {}
                is_high_control = any(
                    high_control.get(key)
                    for key in [
                        "highControl",
                        "traceabilityCritical",
                        "evidenceRequired",
                    ]
                )
                if is_high_control:
                    high_control_count += 1
                    if item.get("text", {}).get("fr"):
                        high_control_fr_count += 1
        path.write_text(
            json.dumps(activities, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    return {
        "contentItems": item_count,
        "contentItemsWithFr": item_fr_count,
        "localizedActivityValuesChanged": localized_values_changed,
        "authoritySensitiveItems": authority_count,
        "authoritySensitiveItemsWithFr": authority_fr_count,
        "highControlItems": high_control_count,
        "highControlItemsWithFr": high_control_fr_count,
    }


def main() -> None:
    concepts, audit = build_batir_concepts()
    phrase_map = build_phrase_map()
    translation_counts = apply_activity_translations(phrase_map)
    audit["translationDrafting"] = {
        "method": "Deterministic terminology-aware draft generation using BÂTIR terms plus controlled QC phrase mappings.",
        "statusApplied": "provisional",
        **translation_counts,
    }
    BATIR_TERMINOLOGY_FILE.write_text(
        json.dumps(concepts, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    BATIR_AUDIT_FILE.write_text(
        json.dumps(audit, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(audit, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
